"""Generate the CCIS Excel dashboard."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analysis import redundancy_matrix, unique_airports_by_card
from src.import_data import airport_coverage_summary, card_airport_counts, load_portfolio_config
from src.lounge_engine import spend_recommendations
from src.spend_tracker import build_milestone_rows, build_spend_tracker_rows

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "dashboard" / "Credit_Card_Intelligence_System.xlsx"

INSTRUCTION_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _autosize_columns(writer_sheet) -> None:
    for column_cells in writer_sheet.columns:
        length = 0
        column = column_cells[0].column
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        writer_sheet.column_dimensions[get_column_letter(column)].width = min(max(length + 2, 12), 48)


def _style_header_row(sheet) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _format_month_label(month_key: str) -> str:
    """Turn 2026-04 into Apr 2026."""
    try:
        year, month = month_key.split("-")
        names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        return f"{names[int(month) - 1]} {year}"
    except (ValueError, IndexError):
        return month_key


def _spend_tracker_dataframe() -> pd.DataFrame:
    rows = build_spend_tracker_rows()
    if not rows:
        return pd.DataFrame()

    months = [rows[0].get("month_1"), rows[0].get("month_2"), rows[0].get("month_3")]
    month_labels = [_format_month_label(m) if m else f"Month {i}" for i, m in enumerate(months, 1)]

    flat_rows = []
    for row in rows:
        flat = {
            "card": row["card_name"],
            "card_key": row["card_key"],
            month_labels[0]: row.get("spend_month_1_inr", 0),
            month_labels[1]: row.get("spend_month_2_inr", 0) if len(month_labels) > 1 else 0,
            month_labels[2]: row.get("spend_month_3_inr", 0) if len(month_labels) > 2 else 0,
            "rolling_3mo_total_inr": row["rolling_3mo_total_inr"],
        }
        flat_rows.append(flat)
    return pd.DataFrame(flat_rows)


def _milestone_dataframe() -> pd.DataFrame:
    rows = build_milestone_rows()
    display = []
    for row in rows:
        display.append(
            {
                "card": row["card_name"],
                "rolling_3mo_spend_inr": row["rolling_3mo_total_inr"],
                "lounge_target_inr": row["lounge_target_inr"],
                "remaining_inr": row["remaining_inr"],
                "pct_to_lounge": row["pct_to_lounge"],
                "lounge_eligible": row["lounge_eligible"],
                "visits_per_quarter": row["lounge_visits_per_quarter"],
            }
        )
    return pd.DataFrame(display)


def _redundancy_dataframe(conn: sqlite3.Connection) -> pd.DataFrame:
    keys, matrix = redundancy_matrix(conn)
    labels = {k: k.replace("_", " ").title()[:20] for k in keys}
    data = {labels[col]: [row[idx] for row in matrix] for idx, col in enumerate(keys)}
    return pd.DataFrame(data, index=[labels[k] for k in keys])


def _add_sheet_note(sheet, text: str, column_count: int) -> None:
    sheet.insert_rows(1)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(column_count, 1))
    cell = sheet.cell(row=1, column=1, value=text)
    cell.fill = INSTRUCTION_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 36


def build_dashboard(conn: sqlite3.Connection, output_path: Path | str | None = None) -> Path:
    output = Path(output_path or DEFAULT_OUTPUT)
    output.parent.mkdir(parents=True, exist_ok=True)
    config = load_portfolio_config()

    cards_df = pd.read_sql_query(
        "SELECT * FROM cards ORDER BY spend_priority_rank IS NULL, spend_priority_rank, card_name",
        conn,
    )
    counts = card_airport_counts(conn)
    cards_df["domestic_airports"] = cards_df["card_key"].map(counts)
    cards_df["lounge_spend_required"] = cards_df["lounge_spend_required"].map({1: "Yes", 0: "No"})

    matrix = pd.DataFrame(airport_coverage_summary(conn))
    card_keys = [card["key"] for card in config["cards"]]
    for key in card_keys:
        if key not in matrix.columns:
            matrix[key] = ""
    matrix = matrix[["airport", *card_keys, "best_free", "best_spend"]]

    spend_df = pd.DataFrame(spend_recommendations(conn))
    frequent = config.get("frequent_airports", [])
    frequent_df = matrix[
        matrix["airport"].isin([a.title() if a.islower() else a for a in frequent])
        | matrix["airport"].isin(frequent)
    ].copy()

    summary_rows = []
    milestone_preview = build_milestone_rows()
    milestone_by_key = {row["card_key"]: row for row in milestone_preview}
    for card in config["cards"]:
        key = card["key"]
        ms = milestone_by_key.get(key, {})
        summary_rows.append(
            {
                "card": card["name"],
                "lounge_eligible": ms.get("lounge_eligible", ""),
                "rolling_3mo_spend_inr": ms.get("rolling_3mo_total_inr", 0),
                "lounge_target_inr": card.get("lounge_spend_inr"),
                "remaining_inr": ms.get("remaining_inr"),
                "domestic_airports": counts.get(key, 0),
                "spend_priority": card.get("spend_priority_rank"),
            }
        )
    summary_df = pd.DataFrame(summary_rows)

    spend_tracker_df = _spend_tracker_dataframe()
    milestone_df = _milestone_dataframe()
    unique_df = pd.DataFrame(unique_airports_by_card(conn, vs_baseline=True))
    redundancy_df = _redundancy_dataframe(conn)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Dashboard", index=False)
        spend_tracker_df.to_excel(writer, sheet_name="Spend Tracker", index=False)
        milestone_df.to_excel(writer, sheet_name="Milestones", index=False)
        cards_df.to_excel(writer, sheet_name="Card Master", index=False)
        matrix.to_excel(writer, sheet_name="Airport Matrix", index=False)
        unique_df.to_excel(writer, sheet_name="Unique Airports", index=False)
        redundancy_df.to_excel(writer, sheet_name="Redundancy", index=True)
        frequent_df.to_excel(writer, sheet_name="Frequent Airports", index=False)
        spend_df.to_excel(writer, sheet_name="Spend Priority", index=False)

        spend_note = (
            "Edit config/spend_tracker.yaml with monthly spends, then run: python main.py build"
        )
        milestone_note = (
            "Lounge eligibility = rolling 3-month spend vs target. "
            "DBS/Tiger always eligible (no spend)."
        )
        unique_note = "Airports each card covers beyond the DBS + Tiger baseline."

        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            _style_header_row(sheet)
            _autosize_columns(sheet)

        if "Spend Tracker" in writer.sheets:
            st = writer.sheets["Spend Tracker"]
            _add_sheet_note(st, spend_note, len(spend_tracker_df.columns))
            _style_header_row(st)

        if "Milestones" in writer.sheets:
            ms = writer.sheets["Milestones"]
            _add_sheet_note(ms, milestone_note, len(milestone_df.columns))
            _style_header_row(ms)

        if "Unique Airports" in writer.sheets:
            ua = writer.sheets["Unique Airports"]
            _add_sheet_note(ua, unique_note, len(unique_df.columns))
            _style_header_row(ua)

    return output
